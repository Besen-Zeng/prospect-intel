"""
main.py — orchestrate the prospect intelligence pipeline.

Reads:   prospects.csv
Writes:  output/report.xlsx

Pipeline per row:
  1. fetcher.py scrapes the company website
  2. analyzer.py sends data to Claude for 5P analysis
  3. Results merged and written to Excel with colour-coded opportunity levels
"""

import csv
import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import fetcher
import analyzer

load_dotenv(override=True)

INPUT_FILE  = "prospects.csv"
OUTPUT_DIR  = "output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "report.xlsx")

# Colour fills for opportunity_level column
FILL_HIGH = PatternFill("solid", fgColor="C6EFCE")   # green
FILL_MID  = PatternFill("solid", fgColor="FFEB9C")   # yellow
FILL_LOW  = PatternFill("solid", fgColor="FFC7CE")   # red
FILL_MAP  = {"High": FILL_HIGH, "Mid": FILL_MID, "Low": FILL_LOW}

# Light blue: client_type cell was blank so auto-inferred value is shown
FILL_AUTO = PatternFill("solid", fgColor="DDEBF7")

# Inserted right after the "client_type" CSV column in the Excel output
CLASSIFICATION_COLS = ["suggested_client_type", "type_reasoning"]

# Appended after all CSV columns
ANALYSIS_COLS = [
    "people", "product", "place", "price",
    "promotion", "opportunity_level", "next_action",
]

# All analysis-derived keys — used to exclude them from the csv_cols list
ALL_ANALYSIS = set(CLASSIFICATION_COLS + ANALYSIS_COLS)


def main():
    rows = _read_csv(INPUT_FILE)
    results = []

    print(f"\nprospect-intel: processing {len(rows)} prospects\n{'─' * 45}")

    for i, row in enumerate(rows, 1):
        company = row.get("company", f"Row {i}")
        print(f"[{i}/{len(rows)}] Researching {company}...")

        site_data = fetcher.fetch(row.get("website", ""), company=company)
        analysis  = analyzer.analyze(row, site_data["text"])

        results.append({**row, **analysis, "_fetch_error": site_data["error"] or ""})

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    saved_path = _write_excel(results)

    print(f"\n{'─' * 45}")
    print(f"Done! Report saved to {saved_path}\n")


def _read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _write_excel(results: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospect Intelligence"

    # Build header list: CSV cols, then classification cols injected after client_type, then 5P cols
    csv_cols = [k for k in results[0].keys() if not k.startswith("_") and k not in ALL_ANALYSIS]
    ct_pos   = csv_cols.index("client_type") + 1 if "client_type" in csv_cols else len(csv_cols)
    headers  = csv_cols[:ct_pos] + CLASSIFICATION_COLS + csv_cols[ct_pos:] + ANALYSIS_COLS

    # Header row — bold, grey background
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font      = Font(bold=True)
        cell.fill      = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Data rows
    opp_col_idx = headers.index("opportunity_level") + 1

    for row_idx, row in enumerate(results, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # If client_type was left blank, display the auto-inferred value (blue tint)
            if key == "client_type" and not row.get("client_type", "").strip():
                cell.value = row.get("suggested_client_type", "")
                cell.fill  = FILL_AUTO

            # Colour the opportunity_level cell
            elif col_idx == opp_col_idx:
                level = row.get("opportunity_level", "")
                if level in FILL_MAP:
                    cell.fill = FILL_MAP[level]

    # Auto-fit column widths (capped at 50 chars)
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(results) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    path = OUTPUT_FILE
    try:
        wb.save(path)
    except PermissionError:
        # File is open in Excel — save a timestamped copy instead
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path  = os.path.join(OUTPUT_DIR, f"report_{stamp}.xlsx")
        wb.save(path)
        print(f"  NOTE: report.xlsx was open — saved as {os.path.basename(path)} instead")
    return path


if __name__ == "__main__":
    main()
